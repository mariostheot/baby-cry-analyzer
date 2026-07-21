#!/usr/bin/env python3
"""
Developer-time tool: download official WHO 2006 Child Growth Standards z-score tables
and write compact JSON assets for the Android app.

Requires: pip install openpyxl

Do NOT run at app runtime.
"""

from __future__ import annotations

import json
import sys
import urllib.request
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "app" / "src" / "main" / "assets" / "who_cgs2006"

WHO_STANDARD = "WHO Child Growth Standards, 2006"

SOURCES = {
    "wfa_boy": {
        "indicator": "weight-for-age",
        "sex": "boy",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/weight-for-age/wfa_boys_0-to-5-years_zscores.xlsx?sfvrsn=97a05331_9",
        "unit": "grams",
        "value_scale": 1000.0,  # official tables are kg -> grams
    },
    "wfa_girl": {
        "indicator": "weight-for-age",
        "sex": "girl",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/weight-for-age/wfa_girls_0-to-5-years_zscores.xlsx?sfvrsn=4c03b8db_7",
        "unit": "grams",
        "value_scale": 1000.0,
    },
    "lhfa_boy_0_2": {
        "indicator": "length-height-for-age",
        "sex": "boy",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/lhfa_boys_0-to-2-years_zscores.xlsx?sfvrsn=30e044c_9",
        "unit": "millimeters",
        "value_scale": 10.0,  # official tables are cm -> mm
        "month_range": (0, 23),
    },
    "lhfa_boy_2_5": {
        "indicator": "length-height-for-age",
        "sex": "boy",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/lhfa_boys_2-to-5-years_zscores.xlsx?sfvrsn=17e5ad91_9",
        "unit": "millimeters",
        "value_scale": 10.0,
        "month_range": (24, 60),
    },
    "lhfa_girl_0_2": {
        "indicator": "length-height-for-age",
        "sex": "girl",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/lhfa_girls_0-to-2-years_zscores.xlsx?sfvrsn=e9e66a95_11",
        "unit": "millimeters",
        "value_scale": 10.0,
        "month_range": (0, 23),
    },
    "lhfa_girl_2_5": {
        "indicator": "length-height-for-age",
        "sex": "girl",
        "url": "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators/length-height-for-age/lhfa_girls_2-to-5-years_zscores.xlsx?sfvrsn=2ec187b9_11",
        "unit": "millimeters",
        "value_scale": 10.0,
        "month_range": (24, 60),
    },
}


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(
            "openpyxl is required to parse WHO XLSX files.\n"
            "Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "baby-cry-analyzer-who-asset-generator/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def _norm_header(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip().lower().replace(" ", "")


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    """Locate the row containing Month and SD columns."""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            key = _norm_header(cell)
            if key in ("month", "μήνας"):
                mapping["month"] = col_idx
            elif key in ("sd2neg", "sd2-", "sd2neg.", "-2sd"):
                mapping["sd2neg"] = col_idx
            elif key in ("sd0", "sd0.", "0sd", "median", "m"):
                if "sd0" not in mapping:
                    mapping["sd0"] = col_idx
            elif key in ("sd2", "sd2+", "sd2.", "+2sd"):
                if "sd2" not in mapping and "sd2neg" in mapping:
                    # avoid matching sd2neg again
                    if col_idx != mapping.get("sd2neg"):
                        mapping["sd2"] = col_idx
                elif "sd2" not in mapping:
                    mapping["sd2"] = col_idx
        # WHO files use explicit SD2neg, SD0, SD2 headers
        for col_idx, cell in enumerate(row):
            raw = str(cell or "").strip()
            low = raw.lower()
            if low == "sd2neg":
                mapping["sd2neg"] = col_idx
            elif low == "sd0":
                mapping["sd0"] = col_idx
            elif low == "sd2" and "sd2neg" not in mapping or col_idx != mapping.get("sd2neg"):
                if low == "sd2" and raw == "SD2":
                    mapping["sd2"] = col_idx
        if all(k in mapping for k in ("month", "sd2neg", "sd0", "sd2")):
            return row_idx, mapping
    raise ValueError("Could not find header row with Month, SD2neg, SD0, SD2")


def parse_xlsx(data: bytes, value_scale: float) -> dict[int, dict[str, int]]:
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb.active
    header_row, cols = find_header_row(sheet)
    by_month: dict[int, dict[str, int]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[cols["month"]] is None:
            continue
        try:
            month = int(float(row[cols["month"]]))
        except (TypeError, ValueError):
            continue
        if month < 0 or month > 60:
            continue

        def to_int(col: str) -> int:
            val = row[cols[col]]
            if val is None:
                raise ValueError(f"Missing {col} at month {month}")
            return int(round(float(val) * value_scale))

        rec = {
            "sd2neg": to_int("sd2neg"),
            "sd0": to_int("sd0"),
            "sd2": to_int("sd2"),
        }
        if month in by_month:
            raise ValueError(f"Duplicate month {month}")
        by_month[month] = rec
    wb.close()
    return by_month


def validate_records(records: list[dict[str, Any]], unit: str) -> None:
    if len(records) != 61:
        raise ValueError(f"Expected 61 records, got {len(records)}")
    for i, rec in enumerate(records):
        if rec["month"] != i:
            raise ValueError(f"Expected month {i}, got {rec['month']}")
        lower, median, upper = rec["sd2neg"], rec["sd0"], rec["sd2"]
        if not (lower < median < upper):
            raise ValueError(f"Month {i}: expected sd2neg < sd0 < sd2, got {lower}, {median}, {upper}")
        if i > 0:
            prev = records[i - 1]["sd0"]
            if median < prev:
                raise ValueError(f"Month {i}: median {median} < previous {prev} (not monotonic)")


def build_series(by_month: dict[int, dict[str, int]]) -> list[dict[str, int]]:
    records = []
    for m in range(61):
        if m not in by_month:
            raise ValueError(f"Missing month {m}")
        records.append({"month": m, **by_month[m]})
    validate_records(records, "")
    return records


def write_asset(path: Path, meta: dict[str, Any], records: list[dict[str, int]]) -> None:
    payload = {**meta, "records": records}
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"Wrote {path} ({len(records)} records)")


def main() -> None:
    _require_openpyxl()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for key in ("wfa_boy", "wfa_girl"):
        cfg = SOURCES[key]
        print(f"Downloading {key}...")
        raw = download(cfg["url"])
        by_month = parse_xlsx(raw, cfg["value_scale"])
        records = build_series(by_month)
        out_name = "wfa_boy.json" if key == "wfa_boy" else "wfa_girl.json"
        write_asset(
            OUT_DIR / out_name,
            {
                "whoStandard": WHO_STANDARD,
                "indicator": cfg["indicator"],
                "sex": cfg["sex"],
                "unit": cfg["unit"],
                "sourceUrl": cfg["url"],
            },
            records,
        )

    for sex in ("boy", "girl"):
        print(f"Merging LHFA {sex}...")
        cfg02 = SOURCES[f"lhfa_{sex}_0_2"]
        cfg25 = SOURCES[f"lhfa_{sex}_2_5"]
        m02 = parse_xlsx(download(cfg02["url"]), cfg02["value_scale"])
        m25 = parse_xlsx(download(cfg25["url"]), cfg25["value_scale"])
        lo, hi = cfg02["month_range"]
        for m in range(lo, hi + 1):
            if m not in m02:
                raise ValueError(f"Missing month {m} in 0-2 file for {sex}")
        lo, hi = cfg25["month_range"]
        for m in range(lo, hi + 1):
            if m not in m25:
                raise ValueError(f"Missing month {m} in 2-5 file for {sex}")
        merged: dict[int, dict[str, int]] = {}
        for m in range(0, 24):
            merged[m] = m02[m]
        for m in range(24, 61):
            merged[m] = m25[m]
        records = build_series(merged)
        write_asset(
            OUT_DIR / f"lhfa_{sex}.json",
            {
                "whoStandard": WHO_STANDARD,
                "indicator": "length-height-for-age",
                "sex": sex,
                "unit": "millimeters",
                "sourceUrl": cfg02["url"],
                "sourceUrlSupplement": cfg25["url"],
            },
            records,
        )

    print("Done.")


if __name__ == "__main__":
    main()
